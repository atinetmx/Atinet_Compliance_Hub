# -*- coding: utf-8 -*-
"""
Script para exportar datos de la BD a Excel
Lee tablas de MySQL y actualiza el archivo Excel preservando formato
"""
import pymysql
import openpyxl
from datetime import datetime

# Configuración BD
CONFIG_BD = {
    'host': 'srvatinet',
    'user': 'root',
    'password': '123456',
    'database': '126ticul20260413',
    'charset': 'utf8'
}

def conectar_bd():
    """Conecta a la base de datos"""
    try:
        conexion = pymysql.connect(**CONFIG_BD)
        print(f"OK: Conectado a {CONFIG_BD['database']}")
        return conexion
    except Exception as e:
        print(f"ERROR al conectar: {e}")
        return None

def exportar_bd_a_excel(archivo_excel_destino):
    """Exporta datos de BD al Excel"""
    
    print("\n" + "="*80)
    print("EXPORTANDO BD A EXCEL")
    print("="*80 + "\n")
    
    # Conectar a BD
    conexion = conectar_bd()
    if not conexion:
        return False
    
    cursor = conexion.cursor()
    
    # Abrir Excel (o crear uno nuevo basado en template)
    try:
        wb = openpyxl.load_workbook(archivo_excel_destino)
        print(f"Excel abierto: {archivo_excel_destino}\n")
    except FileNotFoundError:
        print(f"Creando nuevo archivo Excel: {archivo_excel_destino}\n")
        wb = openpyxl.Workbook()
        # Aquí deberías copiar la estructura del template si es necesario
    
    # 1. ETAPAS <- catactividades
    print("1. Exportando catactividades -> ETAPAS...")
    try:
        cursor.execute("SELECT Actividad, Descripcion, Tipo FROM catactividades ORDER BY Actividad")
        datos = cursor.fetchall()
        
        if 'ETAPAS' in wb.sheetnames:
            ws = wb['ETAPAS']
            # Escribir datos desde fila 3
            for i, registro in enumerate(datos, start=3):
                ws.cell(i, 1).value = registro[0]  # Actividad
                ws.cell(i, 2).value = registro[1]  # Descripcion
            print(f"   OK: {len(datos)} registros escritos")
        else:
            print("   ADVERTENCIA: Hoja ETAPAS no encontrada")
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 2. DOCUMENTOS <- catdocumentos
    print("2. Exportando catdocumentos -> DOCUMENTOS...")
    try:
        cursor.execute("SELECT Documento, Descripcion, Tipo FROM catdocumentos ORDER BY Documento")
        datos = cursor.fetchall()
        
        if 'DOCUMENTOS' in wb.sheetnames:
            ws = wb['DOCUMENTOS']
            for i, registro in enumerate(datos, start=3):
                ws.cell(i, 1).value = registro[0]
                ws.cell(i, 2).value = registro[1]
            print(f"   OK: {len(datos)} registros escritos")
        else:
            print("   ADVERTENCIA: Hoja DOCUMENTOS no encontrada")
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 3. TIPO COMPARECIENTE <- personalidades
    print("3. Exportando personalidades -> TIPO COMPARECIENTE...")
    try:
        cursor.execute("SELECT Id, Descripcion FROM personalidades ORDER BY Id")
        datos = cursor.fetchall()
        
        if 'TIPO COMPARECIENTE' in wb.sheetnames:
            ws = wb['TIPO COMPARECIENTE']
            for i, registro in enumerate(datos, start=3):
                ws.cell(i, 1).value = registro[0]
                ws.cell(i, 2).value = registro[1]
            print(f"   OK: {len(datos)} registros escritos")
        else:
            print("   ADVERTENCIA: Hoja TIPO COMPARECIENTE no encontrada")
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 4. OPERACIONES - AV <- operaciones + opevul
    print("4. Exportando operaciones + opevul -> OPERACIONES...")
    try:
        # Query con JOIN para obtener operación y su actividad vulnerable
        cursor.execute("""
            SELECT 
                o.Operacion,
                o.Descripcion,
                v.Vulnerable
            FROM operaciones o
            LEFT JOIN opevul v ON o.Operacion = v.Operacion
            ORDER BY o.Operacion
        """)
        datos = cursor.fetchall()
        
        # Buscar la hoja de operaciones (puede ser "OPERACIONES - A.V" o "OPERACIONES - AV")
        hoja_operaciones = None
        for nombre in wb.sheetnames:
            if 'OPERACIONES' in nombre:
                hoja_operaciones = nombre
                break
        
        if hoja_operaciones:
            ws = wb[hoja_operaciones]
            for i, registro in enumerate(datos, start=3):
                ws.cell(i, 1).value = registro[0]  # Operacion
                ws.cell(i, 2).value = registro[1]  # Descripcion
                ws.cell(i, 3).value = registro[2]  # NO. AV (Vulnerable)
            print(f"   OK: {len(datos)} registros escritos")
        else:
            print("   ADVERTENCIA: Hoja OPERACIONES - AV no encontrada")
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 5. CONFIG. TRAMITES <- catimpuestos (columnas 1-5)
    print("5. Exportando catimpuestos -> CONFIG. TRAMITES...")
    try:
        cursor.execute("""
            SELECT Impuesto, NomImpuesto, Tipo, cc, Dependencia, cc2 
            FROM catimpuestos 
            ORDER BY Impuesto
        """)
        datos = cursor.fetchall()
        
        # Buscar hoja CONFIG (puede tener variaciones en el nombre)
        hoja_config = None
        for nombre in wb.sheetnames:
            if 'CONFIG' in nombre and ('TRAMITE' in nombre or 'TRÁMITE' in nombre):
                hoja_config = nombre
                break
        
        if hoja_config:
            ws = wb[hoja_config]
            for i, registro in enumerate(datos, start=3):
                ws.cell(i, 1).value = registro[0]  # Impuesto
                ws.cell(i, 2).value = registro[1]  # NomImpuesto
                ws.cell(i, 3).value = registro[2]  # Tipo
                ws.cell(i, 4).value = registro[3]  # cc
                ws.cell(i, 5).value = registro[4]  # Dependencia
            print(f"   OK: {len(datos)} registros escritos")
        else:
            print("   ADVERTENCIA: Hoja CONFIG. TRAMITES no encontrada")
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 6. CONFIG. TRAMITES <- cat_dependencias (columnas 7-8)
    print("6. Exportando cat_dependencias -> CONFIG. TRAMITES...")
    try:
        cursor.execute("SELECT id, Dependencia FROM cat_dependencias ORDER BY id")
        datos = cursor.fetchall()
        
        if hoja_config:
            ws = wb[hoja_config]
            for i, registro in enumerate(datos, start=3):
                ws.cell(i, 7).value = registro[0]  # id
                ws.cell(i, 8).value = registro[1]  # Dependencia
            print(f"   OK: {len(datos)} registros escritos")
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 7. ZONAS <- catzonas (crear hoja si no existe + llenar tabla en COSTOS TRAMITES + TARIFA ISAI)
    print("7. Exportando catzonas -> ZONAS y COSTOS TRAMITES (columnas F-G) + Tarifa ISAI...")
    try:
        cursor.execute("SELECT Zona, Descripcion FROM catzonas ORDER BY Zona")
        datos = cursor.fetchall()
        
        # Obtener tarifas del ISAI (ID=49) por zona (usa PORCENTAJE, no CuotaFijaPesos)
        cursor.execute("""
            SELECT Zona, Porcentaje
            FROM zonastramites
            WHERE Tramite = 49
            ORDER BY Zona
        """)
        tarifas_isai = {z: t for z, t in cursor.fetchall()}  # Diccionario Zona -> Porcentaje ISAI
        
        # A) Crear hoja ZONAS si no existe
        if 'ZONAS' not in wb.sheetnames:
            ws = wb.create_sheet('ZONAS')
            # Crear encabezados
            ws.cell(1, 1).value = None  # Fila 1 vacía
            ws.cell(2, 1).value = "NO. ZONA"
            ws.cell(2, 2).value = "DESCRIPCIÓN ZONA"
            ws.cell(2, 3).value = "TARIFA ISAI"  # Nueva columna
            print("   INFO: Hoja ZONAS creada")
        else:
            ws = wb['ZONAS']
            # Agregar encabezado TARIFA ISAI si no existe
            ws.cell(2, 3).value = "TARIFA ISAI"
        
        # Escribir datos desde fila 3
        for i, registro in enumerate(datos, start=3):
            zona_id = registro[0]
            ws.cell(i, 1).value = zona_id                 # NO. ZONA
            ws.cell(i, 2).value = registro[1]             # DESCRIPCIÓN
            ws.cell(i, 3).value = tarifas_isai.get(zona_id, 0)  # TARIFA ISAI
        print(f"   OK: Hoja ZONAS - {len(datos)} registros escritos (con tarifa ISAI)")
        
        # B) También llenar columnas F-G en COSTOS TRAMITES
        hoja_costos_tram = None
        for nombre in wb.sheetnames:
            if 'COSTOS' in nombre and 'TRAMITE' in nombre:
                hoja_costos_tram = nombre
                break
        
        if hoja_costos_tram:
            ws_ct = wb[hoja_costos_tram]
            # Escribir datos de zonas en columnas F-G desde fila 3
            for i, registro in enumerate(datos, start=3):
                ws_ct.cell(i, 6).value = registro[0]  # F: NO. ZONA
                ws_ct.cell(i, 7).value = registro[1]  # G: NOMBRE ZONA
            print(f"   OK: COSTOS TRAMITES (cols F-G) - {len(datos)} registros escritos")
        
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 8. COSTOS HONORARIOS <- zonasoperaciones
    print("8. Exportando zonasoperaciones -> COSTOS HONORARIOS...")
    try:
        cursor.execute("""
            SELECT 
                zo.Zona,
                zo.Operacion,
                zo.CuotaFijaPesos,
                zo.CuotaFijaUMA,
                zo.Porcentaje,
                zo.Rango,
                zo.RangoP,
                zo.Gastos,
                zo.Administrativo,
                o.Descripcion
            FROM zonasoperaciones zo
            LEFT JOIN operaciones o ON zo.Operacion = o.Operacion
            ORDER BY zo.Zona, zo.Operacion
        """)
        datos = cursor.fetchall()
        
        # Buscar hoja COSTOS HONORARIOS
        hoja_costos_hon = None
        for nombre in wb.sheetnames:
            if 'COSTOS' in nombre and 'HONORARIO' in nombre:
                hoja_costos_hon = nombre
                break
        
        if hoja_costos_hon:
            ws = wb[hoja_costos_hon]
            
            # Deshacer celdas combinadas si existen (para evitar el error MergedCell)
            # Primero hacemos una copia de las celdas combinadas para deshacerlas
            merged_ranges = list(ws.merged_cells.ranges)
            for merged_range in merged_ranges:
                ws.unmerge_cells(str(merged_range))
            
            # Agregar encabezados para campos adicionales después de tabla ARANCEL (columnas L-P)
            ws.cell(2, 12).value = "CUOTA UMA"
            ws.cell(2, 13).value = "PORCENTAJE"
            ws.cell(2, 14).value = "RANGO"
            ws.cell(2, 15).value = "RANGO P"
            ws.cell(2, 16).value = "GASTOS"
            ws.cell(2, 17).value = "ADMINISTRATIVO"
            
            # Escribir datos desde fila 3
            # Columnas A-E: datos principales, F: vacía, G-K: tabla ARANCEL, L-Q: campos adicionales
            for i, registro in enumerate(datos, start=3):
                ws.cell(i, 1).value = registro[1]   # A: NO. OPERACIÓN (Operacion)
                ws.cell(i, 2).value = registro[9]   # B: NOMBRE OPERACIÓN (Descripcion)
                ws.cell(i, 3).value = registro[2]   # C: COSTO (CuotaFijaPesos)
                ws.cell(i, 4).value = None           # D: OBSERVACIONES (vacío por ahora)
                ws.cell(i, 5).value = registro[0]   # E: ZONA / ZONAS (Zona ID)
                # Columnas F vacía, G-K son para tabla ARANCEL (no tocar)
                ws.cell(i, 12).value = registro[3]  # L: CuotaFijaUMA
                ws.cell(i, 13).value = registro[4]  # M: Porcentaje
                ws.cell(i, 14).value = registro[5]  # N: Rango
                ws.cell(i, 15).value = registro[6]  # O: RangoP
                ws.cell(i, 16).value = registro[7]  # P: Gastos
                ws.cell(i, 17).value = registro[8]  # Q: Administrativo
            print(f"   OK: {len(datos)} registros escritos (con campos adicionales)")
        else:
            print(f"   NOTA: Hoja COSTOS HONORARIOS no existe en template (datos en BD: {len(datos)} registros)")
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 9. COSTOS TRAMITES <- zonastramites
    print("9. Exportando zonastramites -> COSTOS TRAMITES...")
    try:
        cursor.execute("""
            SELECT 
                zt.Zona,
                zt.Tramite,
                zt.CuotaFijaPesos,
                zt.CuotaFijaUMA,
                zt.Porcentaje,
                zt.Rango,
                zt.RangoP,
                ci.NomImpuesto
            FROM zonastramites zt
            LEFT JOIN catimpuestos ci ON zt.Tramite = ci.Impuesto
            ORDER BY zt.Zona, zt.Tramite
        """)
        datos = cursor.fetchall()
        
        # Buscar hoja COSTOS TRAMITES
        hoja_costos_tram = None
        for nombre in wb.sheetnames:
            if 'COSTOS' in nombre and 'TRAMITE' in nombre:
                hoja_costos_tram = nombre
                break
        
        if hoja_costos_tram:
            ws = wb[hoja_costos_tram]
            
            # Deshacer celdas combinadas si existen
            merged_ranges = list(ws.merged_cells.ranges)
            for merged_range in merged_ranges:
                ws.unmerge_cells(str(merged_range))
            
            # Agregar encabezados para campos adicionales después de tabla ZONAS (columnas H-K)
            ws.cell(2, 8).value = "CUOTA UMA"
            ws.cell(2, 9).value = "PORCENTAJE"
            ws.cell(2, 10).value = "RANGO"
            ws.cell(2, 11).value = "RANGO P"
            
            # Escribir datos desde fila 3
            # Columnas A-D: datos principales, E: vacía, F-G: tabla ZONAS, H-K: campos adicionales
            for i, registro in enumerate(datos, start=3):
                ws.cell(i, 1).value = registro[1]   # A: NO. OPERACIÓN (Tramite ID)
                ws.cell(i, 2).value = registro[7]   # B: NOMBRE OPERACIÓN (NomImpuesto)
                ws.cell(i, 3).value = registro[2]   # C: COSTO (CuotaFijaPesos)
                ws.cell(i, 4).value = registro[0]   # D: ZONA / ZONAS (Zona ID)
                # Columna E vacía (no escribir nada)
                # Columnas F-G NO se tocan (son para tabla ZONAS)
                ws.cell(i, 8).value = registro[3]   # H: CuotaFijaUMA
                ws.cell(i, 9).value = registro[4]   # I: Porcentaje
                ws.cell(i, 10).value = registro[5]  # J: Rango
                ws.cell(i, 11).value = registro[6]  # K: RangoP
            print(f"   OK: {len(datos)} registros escritos (con campos adicionales)")
        else:
            print(f"   NOTA: Hoja COSTOS TRAMITES no encontrada (datos en BD: {len(datos)} registros)")
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 10. Valor UMA <- datosnotario
    print("10. Exportando valor UMA desde datosnotario...")
    try:
        cursor.execute("SELECT Uma, IVA, ISR, RETIVA FROM datosnotario LIMIT 1")
        datos = cursor.fetchone()
        
        if datos:
            # Buscar hoja COSTOS HONORARIOS para actualizar el valor UMA (generalmente en F1)
            if hoja_costos_hon:
                ws = wb[hoja_costos_hon]
                # El valor UMA suele estar en F1 o cerca
                ws.cell(1, 6).value = datos[0]  # Uma en F1
                print(f"   OK: UMA actualizado a {datos[0]}")
            else:
                print(f"   NOTA: Valor UMA: {datos[0]} (no hay hoja donde escribirlo)")
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 11. ARANCEL <- arancel (crear hoja si no existe + llenar tabla en COSTOS HONORARIOS)
    print("11. Exportando tabla arancel -> ARANCEL y COSTOS HONORARIOS (columnas G-K)...")
    try:
        cursor.execute("""
            SELECT Id, LimInf, LimSup, Porcentaje, Descripcion, DiasInf, DiasSup
            FROM arancel
            ORDER BY Id
        """)
        datos = cursor.fetchall()
        
        if len(datos) > 0:
            # A) Crear hoja ARANCEL si no existe
            if 'ARANCEL' not in wb.sheetnames:
                ws = wb.create_sheet('ARANCEL')
                # Crear encabezados
                ws.cell(1, 1).value = None  # Fila 1 vacía
                ws.cell(2, 1).value = "ID"
                ws.cell(2, 2).value = "LÍMITE INFERIOR"
                ws.cell(2, 3).value = "LÍMITE SUPERIOR"
                ws.cell(2, 4).value = "PORCENTAJE"
                ws.cell(2, 5).value = "DESCRIPCIÓN"
                ws.cell(2, 6).value = "DÍAS INFERIOR"
                ws.cell(2, 7).value = "DÍAS SUPERIOR"
                print("   INFO: Hoja ARANCEL creada")
            else:
                ws = wb['ARANCEL']
            
            # Escribir datos desde fila 3
            for i, registro in enumerate(datos, start=3):
                ws.cell(i, 1).value = registro[0]  # Id
                ws.cell(i, 2).value = registro[1]  # LimInf
                ws.cell(i, 3).value = registro[2]  # LimSup
                ws.cell(i, 4).value = registro[3]  # Porcentaje
                ws.cell(i, 5).value = registro[4]  # Descripcion
                ws.cell(i, 6).value = registro[5]  # DiasInf
                ws.cell(i, 7).value = registro[6]  # DiasSup
            print(f"   OK: Hoja ARANCEL - {len(datos)} registros escritos")
            
            # B) También llenar columnas G-K en COSTOS HONORARIOS
            # Buscar hoja COSTOS HONORARIOS
            hoja_costos_hon = None
            for nombre in wb.sheetnames:
                if 'COSTOS' in nombre and 'HONORARIO' in nombre:
                    hoja_costos_hon = nombre
                    break
            
            if hoja_costos_hon:
                ws_ch = wb[hoja_costos_hon]
                # Escribir datos de arancel en columnas G-K desde fila 3
                # Nota: Solo 5 columnas (G=LimInf, H=LimSup, I=Cuota, J=Porcentaje, K=Descripcion)
                for i, registro in enumerate(datos, start=3):
                    ws_ch.cell(i, 7).value = registro[1]   # G: LimInf
                    ws_ch.cell(i, 8).value = registro[2]   # H: LimSup
                    ws_ch.cell(i, 9).value = None          # I: Cuota Pesos (no está en arancel, solo en template como ejemplo)
                    ws_ch.cell(i, 10).value = registro[3]  # J: Porcentaje
                    ws_ch.cell(i, 11).value = registro[4]  # K: Descripcion
                print(f"   OK: COSTOS HONORARIOS (cols G-K) - {len(datos)} registros escritos")
        else:
            print("   NOTA: No hay registros de aranceles en BD")
    except Exception as e:
        print(f"   ERROR: {e}")
    
    # 12. Crear hoja TARIFAS POR ZONA con filtro automático
    print("12. Creando hoja TARIFAS POR ZONA con filtros...")
    try:
        # Crear hoja
        if 'TARIFAS POR ZONA' in wb.sheetnames:
            del wb['TARIFAS POR ZONA']
        ws_tarifas = wb.create_sheet('TARIFAS POR ZONA', 0)  # Primera hoja
        
        # Encabezados
        ws_tarifas.cell(1, 1).value = "ZONA"
        ws_tarifas.cell(1, 2).value = "NOMBRE ZONA"
        ws_tarifas.cell(1, 3).value = "DESCRIPCIÓN"
        ws_tarifas.cell(1, 4).value = "CUOTA FIJA $"
        ws_tarifas.cell(1, 5).value = "CUOTA FIJA UMA"
        ws_tarifas.cell(1, 6).value = "NO. SALARIOS"
        ws_tarifas.cell(1, 7).value = "% IMPUESTO EXTRA"
        ws_tarifas.cell(1, 8).value = "% PORCENTAJE"
        ws_tarifas.cell(1, 9).value = "TARIFA POR RANGO"
        
        # Formato encabezados
        for col in range(1, 10):
            cell = ws_tarifas.cell(1, col)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        
        # Obtener todas las tarifas ordenadas por zona
        cursor.execute("""
            SELECT 
                zt.Zona,
                cz.Descripcion,
                ci.NomImpuesto,
                zt.CuotaFijaPesos,
                zt.CuotaFijaUMA,
                0,  -- No. Salarios (no existe en BD)
                0,  -- % Impuesto Extra (no existe en BD)
                zt.Porcentaje,
                zt.Rango
            FROM zonastramites zt
            LEFT JOIN catzonas cz ON zt.Zona = cz.Zona
            LEFT JOIN catimpuestos ci ON zt.Tramite = ci.Impuesto
            ORDER BY cz.Descripcion, ci.NomImpuesto
        """)
        todas_tarifas = cursor.fetchall()
        
        # Escribir datos
        for i, registro in enumerate(todas_tarifas, start=2):
            ws_tarifas.cell(i, 1).value = registro[0]  # Zona
            ws_tarifas.cell(i, 2).value = registro[1]  # Nombre Zona
            ws_tarifas.cell(i, 3).value = registro[2]  # Descripción
            ws_tarifas.cell(i, 4).value = registro[3]  # Cuota Fija $
            ws_tarifas.cell(i, 5).value = registro[4]  # Cuota UMA
            ws_tarifas.cell(i, 6).value = registro[5]  # No. Salarios
            ws_tarifas.cell(i, 7).value = registro[6]  # % Imp Extra
            ws_tarifas.cell(i, 8).value = registro[7]  # % Porcentaje
            ws_tarifas.cell(i, 9).value = registro[8]  # Tarifa Rango
        
        # Agregar filtro automático
        ws_tarifas.auto_filter.ref = f"A1:I{len(todas_tarifas) + 1}"
        
        # Inmovilizar panel (freeze) en fila 1
        ws_tarifas.freeze_panes = "A2"
        
        # Ajustar anchos de columna
        ws_tarifas.column_dimensions['A'].width = 8
        ws_tarifas.column_dimensions['B'].width = 25
        ws_tarifas.column_dimensions['C'].width = 50
        ws_tarifas.column_dimensions['D'].width = 15
        ws_tarifas.column_dimensions['E'].width = 15
        ws_tarifas.column_dimensions['F'].width = 15
        ws_tarifas.column_dimensions['G'].width = 18
        ws_tarifas.column_dimensions['H'].width = 15
        ws_tarifas.column_dimensions['I'].width = 20
        
        print(f"   OK: Hoja TARIFAS POR ZONA creada con {len(todas_tarifas)} registros")
        print(f"   INFO: Use los filtros en la columna 'NOMBRE ZONA' para seleccionar municipio")
        
    except Exception as e:
        print(f"   ERROR al crear tarifas por zona: {e}")
    
    # Cerrar conexión
    cursor.close()
    conexion.close()
    
    # Guardar Excel
    try:
        wb.save(archivo_excel_destino)
        print("\n" + "="*80)
        print(f"EXCEL ACTUALIZADO: {archivo_excel_destino}")
        print("="*80)
        return True
    except Exception as e:
        print(f"\nERROR al guardar Excel: {e}")
        return False

if __name__ == "__main__":
    # Nombre del archivo Excel a actualizar
    archivo_excel = "CATALOGOS_COMPLETO_FINAL_126TICUL.xlsx"
    
    # Usar el template SIN QUERYS (tiene las hojas de costos limpias)
    try:
        import shutil
        template = "CATALOGOS Y COSTOS ATINET, NOTARIA 99 NOGALES.xlsx"
        shutil.copy(template, archivo_excel)
        print(f"Template copiado: {template} -> {archivo_excel}")
    except:
        print("No se pudo copiar template, se creara nuevo Excel")
    
    # Exportar datos
    resultado = exportar_bd_a_excel(archivo_excel)
    
    if resultado:
        print("\nProceso completado exitosamente!")
        print(f"Archivo: {archivo_excel}")
    else:
        print("\nHubo errores en el proceso")
